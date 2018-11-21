import openpyxl
import xlrd
from collections import OrderedDict
import json
import codecs
from datetime import datetime
from xlrd import xldate_as_tuple


def list_append(convert_list,sh,rows,cols):
    for i in range(rows):
        row_content = []
        for j in range(cols):
            # ctype： 0,empty, 1,string, 2,number, 3,date, 4,boolean, 5,error
            ctype = sh.cell(i, j).ctype  # 表格的数据类型
            # print(sh.cell(i, j),ctype)
            cell = sh.cell_value(i, j)
            if ctype == 2 and cell % 1 == 0:  # 如果是整形
                cell = int(cell)
            elif ctype == 3:
                # 转成datetime对象
                date = datetime(*xldate_as_tuple(cell, 0))
                cell = date.strftime('%Y/%d/%m %H:%M:%S')
            elif ctype == 4:
                cell = True if cell == 1 else False
            row_content.append(cell)
        # print(row_content)
        single = OrderedDict()
        single[row_content[0]] = row_content[1]
        convert_list.append(single)
        # print(convert_list)
    excle = json.dumps(convert_list)
    excle_yangben = json.loads(excle)

    aa=list_to_dict(excle_yangben)
    return aa

def list_to_dict(excle_yangben):
    aa = OrderedDict()
    bb = []
    for i in range(len(excle_yangben)):
        for k, v in excle_yangben[i].items():
            aa[k.lower()] = excle_yangben[i][k]
    bb.append(aa)
    cc = json.dumps(bb)
    dd = json.loads(cc)[0]
    return dd

def get_excel(excel_path):
    wb = xlrd.open_workbook(excel_path)
    # 获取workbook中所有的表格
    sheets = wb.sheet_names()
    convert_list = []
    # 循环遍历所有sheet
    for i in range(len(sheets)):
        sh = wb.sheet_by_index(i)
        rows = sh.nrows
        cols = sh.ncols
        excle_yangben = list_append(convert_list,sh,rows,cols)
    #print("xxx",len(excle_yangben))
    return excle_yangben


def get_text_excel(excle_path="test.xlsx"):
    wb = xlrd.open_workbook(excle_path)
    convert_list = []
    #通过索引获取表格
    sh = wb.sheet_by_index(0)
    #获取表内容
    title = sh.row_values(0)
    #print(title)
    keys = []
    # 获取表行数
    rows=sh.nrows
    # print(sh.nrows)
    # 获取表列数
    cols = sh.ncols
    for i in range(rows):
        row_content = []
        for j in range(cols):
            #ctype： 0,empty, 1,string, 2,number, 3,date, 4,boolean, 5,error
            ctype = sh.cell(i, j).ctype  # 表格的数据类型
            #print(sh.cell(i, j),ctype)
            cell = sh.cell_value(i, j)
            if ctype == 2 and cell % 1 == 0:  # 如果是整形
                cell = int(cell)
            elif ctype == 3:
                # 转成datetime对象
                date = datetime(*xldate_as_tuple(cell, 0))
                cell = date.strftime('%Y/%d/%m %H:%M:%S')
            elif ctype == 4:
                cell = True if cell == 1 else False
            row_content.append(cell)
    #根据按照类型转化后的字典重组
    single = OrderedDict()
    for colnum in range(0, len(row_content)):
        single[title[colnum]] = row_content[colnum]
    convert_list.append(single)
    excle = json.dumps(convert_list)
    excle_yangben = json.loads(excle)
    return excle_yangben[0]

def Comparison_check(aa,bb):
    for key in bb:
        try:
            #相等的值不打印
            if str(bb[key]) in str(aa[key]):
                pass
                # print('变量名：%s，excel样本中的值："%s"，校验值excel中的值："%s"'%(key, bb[key], aa[key]))
            else:
                print('变量名：%s，excel样本中的值："%s"，校验值excel中的值："%s"'%(key, bb[key], aa[key]))

         #打印出在excel中存在，在需要检验的excel中不存在的变量
        except Exception as e:
            pass
            #print('在校验excel文件中不存在%s变量' %e)


if __name__=="__main__":
    Comparison_check(get_excel("jiaoyanbiao.xlsx"),get_text_excel("test.xlsx"))